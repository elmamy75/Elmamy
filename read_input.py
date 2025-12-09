"""
Script pour lire le fichier Input.xlsx
"""
import pandas as pd
from openpyxl import load_workbook


def read_input(filename='Input.xlsx'):
    """
    Lit le fichier Input.xlsx et extrait les données des différentes sections
    
    Args:
        filename: nom du fichier Excel à lire
        
    Returns:
        dict: Dictionnaire contenant les DataFrames de chaque section
    """
    
    # Charger le fichier Excel
    wb = load_workbook(filename)
    ws = wb['INPUT']
    
    # Dictionnaire pour stocker les résultats
    data = {}
    
    # ========== Lecture des Matériaux ==========
    print("Lecture de la section Matériaux...")
    
    # Trouver la ligne de début des matériaux (chercher "Matériaux")
    mat_start_row = None
    for row in range(1, ws.max_row + 1):
        cell_value = ws.cell(row=row, column=2).value
        if cell_value == 'Matériaux':
            mat_start_row = row + 2  # +2 pour passer le titre et arriver aux en-têtes
            break
    
    if mat_start_row:
        # Lire les en-têtes
        headers_mat = []
        col = 2
        while ws.cell(row=mat_start_row, column=col).value:
            headers_mat.append(ws.cell(row=mat_start_row, column=col).value)
            col += 1
        
        # Lire les données
        materials_data = []
        row = mat_start_row + 1
        while ws.cell(row=row, column=2).value is not None:
            row_data = []
            for c in range(2, 2 + len(headers_mat)):
                row_data.append(ws.cell(row=row, column=c).value)
            materials_data.append(row_data)
            row += 1
            # S'arrêter si on trouve une ligne vide
            if ws.cell(row=row, column=2).value is None:
                break
        
        data['Matériaux'] = pd.DataFrame(materials_data, columns=headers_mat)
        print(f"  {len(data['Matériaux'])} matériaux lus")
    
    # ========== Lecture des Membres ==========
    print("\nLecture de la section Membres...")
    
    # Trouver la ligne de début des membres
    mem_start_row = None
    for row in range(1, ws.max_row + 1):
        cell_value = ws.cell(row=row, column=2).value
        if cell_value == 'Membres':
            mem_start_row = row + 2
            break
    
    if mem_start_row:
        # Lire les en-têtes
        headers_mem = []
        col = 2
        while ws.cell(row=mem_start_row, column=col).value:
            headers_mem.append(ws.cell(row=mem_start_row, column=col).value)
            col += 1
        
        # Lire les données
        members_data = []
        row = mem_start_row + 1
        while ws.cell(row=row, column=2).value is not None:
            row_data = []
            for c in range(2, 2 + len(headers_mem)):
                row_data.append(ws.cell(row=row, column=c).value)
            members_data.append(row_data)
            row += 1
            # S'arrêter si on trouve une ligne vide ou le titre suivant
            if ws.cell(row=row, column=2).value is None or ws.cell(row=row, column=2).value in ['Combinaisons analysées']:
                break
        
        data['Membres'] = pd.DataFrame(members_data, columns=headers_mem)
        print(f"  {len(data['Membres'])} membres lus")
    
    # ========== Lecture des Combinaisons analysées ==========
    print("\nLecture de la section Combinaisons analysées...")
    
    # Trouver la ligne de début des combinaisons
    comb_start_row = None
    for row in range(1, ws.max_row + 1):
        cell_value = ws.cell(row=row, column=2).value
        if cell_value == 'Combinaisons analysées':
            comb_start_row = row + 2
            break
    
    if comb_start_row:
        # Lire les données des combinaisons
        combinations_data = []
        row = comb_start_row
        while ws.cell(row=row, column=2).value is not None:
            row_data = []
            col = 2
            while ws.cell(row=row, column=col).value:
                row_data.append(ws.cell(row=row, column=col).value)
                col += 1
            if row_data:
                combinations_data.append(row_data)
            row += 1
            # S'arrêter si on trouve le titre suivant
            if ws.cell(row=row, column=2).value in ['Section']:
                break
        
        # Créer un DataFrame avec des colonnes génériques
        max_cols = max(len(row) for row in combinations_data) if combinations_data else 0
        headers_comb = ['Situation'] + [f'CO{i}' for i in range(1, max_cols)]
        
        # Compléter les lignes avec None si nécessaire
        for row in combinations_data:
            while len(row) < max_cols:
                row.append(None)
        
        data['Combinaisons'] = pd.DataFrame(combinations_data, columns=headers_comb[:max_cols])
        print(f"  {len(data['Combinaisons'])} situations lues")
    
    # ========== Lecture des Sections ==========
    print("\nLecture de la section Sections...")
    
    # Trouver la ligne de début des sections
    sec_start_row = None
    for row in range(1, ws.max_row + 1):
        cell_value = ws.cell(row=row, column=2).value
        if cell_value == 'Section':
            sec_start_row = row + 2
            break
    
    if sec_start_row:
        # Lire les en-têtes
        headers_sec = []
        col = 2
        while ws.cell(row=sec_start_row, column=col).value:
            headers_sec.append(ws.cell(row=sec_start_row, column=col).value)
            col += 1
        
        # Lire les données
        sections_data = []
        row = sec_start_row + 1
        while row <= ws.max_row and ws.cell(row=row, column=2).value is not None:
            row_data = []
            for c in range(2, 2 + len(headers_sec)):
                row_data.append(ws.cell(row=row, column=c).value)
            if any(val is not None for val in row_data):  # Vérifier qu'il y a des données
                sections_data.append(row_data)
            row += 1
        
        data['Sections'] = pd.DataFrame(sections_data, columns=headers_sec)
        print(f"  {len(data['Sections'])} sections lues")
    
    wb.close()
    
    return data


# ========== TEST ==========

if __name__ == "__main__":
    
    try:
        # Lire le fichier Input.xlsx
        print("="*60)
        print("Lecture du fichier Input.xlsx")
        print("="*60)
        
        donnees = read_input('Input.xlsx')
        
        print("\n" + "="*60)
        print("RÉSUMÉ DES DONNÉES LUES")
        print("="*60)
        
        # Afficher les matériaux
        if 'Matériaux' in donnees:
            print("\n--- MATÉRIAUX ---")
            print(donnees['Matériaux'])
        
        # Afficher les membres
        if 'Membres' in donnees:
            print("\n--- MEMBRES ---")
            print(donnees['Membres'])
        
        # Afficher les combinaisons
        if 'Combinaisons' in donnees:
            print("\n--- COMBINAISONS ANALYSÉES ---")
            print(donnees['Combinaisons'])
        
        # Afficher les sections
        if 'Sections' in donnees:
            print("\n--- SECTIONS ---")
            print(donnees['Sections'])
        
        print("\n" + "="*60)
        print("Lecture terminée avec succès !")
        print("="*60)
        
    except FileNotFoundError:
        print("Erreur : Le fichier 'Input.xlsx' n'a pas été trouvé.")
    except Exception as e:
        print(f"Erreur lors de la lecture : {e}")
        import traceback
        traceback.print_exc()
