"""
Créer un fichier excel Input.xlsx formaté selon le modèle
avec les sections : Matériaux, Membres, Combinaisons analysées, et Sections
"""
import pandas as pd
import numpy as np
from openpyxl.styles import Alignment, Border, Side, Font
from openpyxl import load_workbook
from material import Material
from element import Element
from section import Section


def load_sections_from_bdd(bdd_file='BDD_Sections.xlsx'):
    """
    Charge les sections depuis le fichier BDD_Sections.xlsx
    
    Returns:
        dict: Dictionnaire {id: Section} des sections chargées
    """
    # Lire la feuille des caractéristiques géométriques
    df = pd.read_excel(bdd_file, sheet_name='Caractéristiques géométriques', skiprows=3)
    
    # Lire la feuille des stress points
    df_sp = pd.read_excel(bdd_file, sheet_name='Stress Points', skiprows=3)
    
    sections_dict = {}
    section_id = 1
    
    for idx, row in df.iterrows():
        section_name = row['Section']
        
        # Ignorer les lignes vides ou les titres
        if pd.isna(section_name) or section_name in ['Symbol RSTAB', 'Symbol RCC-M', 'Unités']:
            continue
        
        # Vérifier si c'est une section fermée (1) ou ouverte (0)
        is_closed = bool(row['Paroi']) if not pd.isna(row['Paroi']) else False
        
        # Extraire les stress points pour cette section
        stress_points = []
        section_sp = df_sp[df_sp['Section'] == section_name]
        
        for sp_idx, sp_row in section_sp.iterrows():
            if not pd.isna(sp_row['No.']):
                try:
                    stress_point = Section.StressPoint(
                        sec_name=section_name,
                        id=int(sp_row['No.']),
                        coordinates=(float(sp_row['y [mm]']), float(sp_row['z [mm]'])),
                        Qy=float(sp_row['Qy [mm3]']) if not pd.isna(sp_row['Qy [mm3]']) else 0.0,
                        Qz=float(sp_row['Qz [mm3]']) if not pd.isna(sp_row['Qz [mm3]']) else 0.0,
                        e=float(sp_row['t [mm]']) if not pd.isna(sp_row['t [mm]']) else 0.0,
                        Wno=float(sp_row['Wno [mm2]']) if not pd.isna(sp_row['Wno [mm2]']) else 0.0,
                        Sw=float(sp_row['Sv [mm4]']) if not pd.isna(sp_row['Sv [mm4]']) else 0.0
                    )
                    stress_points.append(stress_point)
                except:
                    continue
        
        # Créer l'objet Section
        try:
            section = Section(
                name=section_name,
                is_closed=is_closed,
                h=float(row['Depth']) if not pd.isna(row['Depth']) and row['Depth'] != '-' else 0.0,
                l=float(row['Width']) if not pd.isna(row['Width']) and row['Width'] != '-' else 0.0,
                D=float(row['Diameter']) if not pd.isna(row['Diameter']) and row['Diameter'] != '-' else None,
                tw=float(row['Web thickness']) if not pd.isna(row['Web thickness']) else 0.0,
                tf=float(row['Flange Thickness']) if not pd.isna(row['Flange Thickness']) else 0.0,
                A=float(row['Cross-section area']) if not pd.isna(row['Cross-section area']) else 0.0,
                Iy=float(row['Moment of inertia']) if not pd.isna(row['Moment of inertia']) else 0.0,
                Iz=float(row['Product second moment of area']) if not pd.isna(row['Product second moment of area']) else 0.0,
                ry=float(row['Governing radius of gyration y']) if not pd.isna(row['Governing radius of gyration y']) else 0.0,
                rz=float(row['Governing radius of gyration z']) if not pd.isna(row['Governing radius of gyration z']) else 0.0,
                Am=float(row['Core area']) if not pd.isna(row['Core area']) and row['Core area'] != '-' else 0.0,
                b_t=float(row['b/t au sens du RCC-M']) if not pd.isna(row['b/t au sens du RCC-M']) else 0.0,
                d_t=float(row['d/t au sens du RCC-M']) if not pd.isna(row['d/t au sens du RCC-M']) else 0.0,
                Sp=stress_points
            )
            sections_dict[section_id] = section
            section_id += 1
        except Exception as e:
            print(f"  Erreur lors de la création de la section {section_name}: {e}")
            continue
    
    return sections_dict


def create_input(materials, member, sections, combinations=None):
    """
    Cette fonction lit les dictionnaires materials, member, sections et combinations 
    et va ensuite les mettre en forme pour les intégrer à un fichier excel .xlsx au format souhaité
    
    Args:
        materials: dictionnaire d'objets Material
        member: dictionnaire d'objets Element
        sections: dictionnaire d'objets Section
        combinations: dictionnaire optionnel des combinaisons analysées
    """
    
    # ========== Préparation du tableau des matériaux ==========
    
    df_materials = pd.DataFrame.from_dict(
        {k: vars(v) for k, v in materials.items()},
        orient='index'
    )
  
    df_materials.rename(
        columns={
            "name": "Nom",
            "temperature": "Température [°C]",
            "E": "E [MPa] ",
            "Sy": "Sy [MPa]  ",
            "Su": "Su [MPa] ",
            "poisson": "Coef de Poisson"
        },
        inplace=True
    )
    
    df_materials.insert(0, "ID", df_materials.index)
    
    # ========== Préparation du tableau des membres ==========
    
    df_member = pd.DataFrame.from_dict(
        {k: vars(v) for k, v in member.items()},
        orient='index'
    )
  
    df_member.rename(
        columns={
            "id": "ID",
            "section": "Section",
            "material": "Matériau",
            "lambda_rccm": "Longueur λ [mm]",
            "Lb": "Longueur Lc [mm]"
        },
        inplace=True
    )  
    
    # Ajouter les colonnes manquantes
    df_member["ky"] = 2
    df_member["kz"] = 2
    df_member["Cmy"] = 0.85
    df_member["Cmz"] = 0.85
   
    # Extraire les nœuds de début et fin
    df_member["Nœud début"] = df_member["nodes_id"].apply(
        lambda x: x[0] if isinstance(x, list) and len(x) > 0 else None
    )
    df_member["Nœud fin"] = df_member["nodes_id"].apply(
        lambda x: x[-1] if isinstance(x, list) and len(x) > 0 else None
    )
    
    # Supprimer les colonnes non utiles
    columns_to_drop = ['name', 'nodes_id']
    for col in columns_to_drop:
        if col in df_member.columns:
            df_member.drop(columns=[col], inplace=True)
   
    # Réorganiser les colonnes
    col = df_member.pop("Nœud début")
    df_member.insert(1, "Nœud début", col) 
    col = df_member.pop("Nœud fin")
    df_member.insert(2, "Nœud fin", col)
    
    # ========== Préparation du tableau des sections ==========
    
    # Convertir les objets Section (dataclass) en dictionnaire
    sections_data = {}
    for k, v in sections.items():
        section_dict = {
            'name': v.name,
            'h': v.h,
            'l': v.l,  # Utiliser 'l' au lieu de 'b'
            'tw': v.tw,
            'tf': v.tf,
            'A': v.A,
            'Iy': v.Iy,
            'Iz': v.Iz,
            'ry': v.ry,
            'rz': v.rz,
            'J': 0,  # Vous pouvez ajouter J si disponible
            'Ac': v.Am if v.Am != 0.0 else '-',
            'Qy': v.Sp[0].Qy if v.Sp else 0,  # Premier stress point
            'Qz': v.Sp[0].Qz if v.Sp else 0
        }
        sections_data[k] = section_dict
    
    df_sections = pd.DataFrame.from_dict(sections_data, orient='index')
    
    # Renommer les colonnes pour les sections
    df_sections.rename(
        columns={
            "name": "Nom",
            "h": "h [mm]",
            "l": "l [mm]",
            "tw": "tw [mm]",
            "tf": "tf [mm]",
            "A": "A [mm²]",
            "Iy": "Iy [mm4]",
            "Iz": "Iz [mm4]",
            "ry": "ry [mm]",
            "rz": "rz [mm]",
            "J": "J [mm]",
            "Ac": "Ac [mm²]",
            "Qy": "Qy [mm3]",
            "Qz": "Qz [mm3]"
        },
        inplace=True
    )
    
    df_sections.insert(0, "ID", df_sections.index)
    
    # ========== Création du fichier Excel ==========
    
    nom_fichier = "Input.xlsx"
    
    # Créer un fichier Excel vide
    with pd.ExcelWriter(nom_fichier, engine="openpyxl") as writer:
        workbook = writer.book
        worksheet = workbook.create_sheet('INPUT', 0)
        
        if 'Sheet' in workbook.sheetnames:
            workbook.remove(workbook['Sheet'])
    
    # Recharger le fichier pour ajouter le contenu
    wb = load_workbook(nom_fichier)
    ws = wb['INPUT']
    
    # ========== Ajouter le texte d'en-tête ==========
    ws['B2'] = ('Les données ci-dessous ont été extraites de la mise en donnée logiciel. \n'
                'Vérifiez que les données extraites correspondent bien à ce qui est attendu.\n'
                'Nota : les valeurs de "k" et de "Cm" peuvent  être modifiées pour optimiser le calcul.')
    ws['B2'].alignment = Alignment(wrap_text=True, vertical='top')
    ws.row_dimensions[2].height = 45
    
    # ========== Section Matériaux ==========
    
    current_row = 9
    
    # Titre "Matériaux"
    ws.cell(row=current_row, column=2, value='Matériaux')
    ws.cell(row=current_row, column=2).font = Font(bold=True, size=12)
    
    current_row += 2  # Ligne 11
    
    # En-têtes des colonnes pour les matériaux
    headers_materials = list(df_materials.columns)
    for col_idx, header in enumerate(headers_materials, start=2):
        cell = ws.cell(row=current_row, column=col_idx, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    
    current_row += 1
    
    # Données des matériaux
    for row_data in df_materials.itertuples(index=False):
        for col_idx, value in enumerate(row_data, start=2):
            cell = ws.cell(row=current_row, column=col_idx, value=value)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
        current_row += 1
    
    # ========== Section Membres ==========
    
    current_row += 3  # Espacement
    
    # Titre "Membres"
    ws.cell(row=current_row, column=2, value='Membres')
    ws.cell(row=current_row, column=2).font = Font(bold=True, size=12)
    
    current_row += 2
    
    # En-têtes des colonnes pour les membres
    headers_members = list(df_member.columns)
    for col_idx, header in enumerate(headers_members, start=2):
        cell = ws.cell(row=current_row, column=col_idx, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    
    current_row += 1
    
    # Données des membres
    for row_data in df_member.itertuples(index=False):
        for col_idx, value in enumerate(row_data, start=2):
            cell = ws.cell(row=current_row, column=col_idx, value=value)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
        current_row += 1
    
    # ========== Section Combinaisons analysées ==========
    
    current_row += 4  # Espacement
    
    # Titre "Combinaisons analysées"
    ws.cell(row=current_row, column=2, value='Combinaisons analysées')
    ws.cell(row=current_row, column=2).font = Font(bold=True, size=12)
    
    current_row += 2
    
    # Si des combinaisons sont fournies
    if combinations:
        for situation, combos in combinations.items():
            row_data = [situation] + combos
            for col_idx, value in enumerate(row_data, start=2):
                cell = ws.cell(row=current_row, column=col_idx, value=value)
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
            current_row += 1
    else:
        # Valeurs par défaut
        default_combos = [
            ['Situation 1', 'NP', 'CO3', 'CO4', 'CO5'],
            ['Situation 2', 'ACC', 'CO9', 'CO10', 'CO11']
        ]
        for row_data in default_combos:
            for col_idx, value in enumerate(row_data, start=2):
                cell = ws.cell(row=current_row, column=col_idx, value=value)
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
            current_row += 1
    
    # ========== Section Sections ==========
    
    current_row += 3  # Espacement
    
    # Titre "Section"
    ws.cell(row=current_row, column=2, value='Section')
    ws.cell(row=current_row, column=2).font = Font(bold=True, size=12)
    
    current_row += 2
    
    # En-têtes des colonnes pour les sections
    headers_sections = list(df_sections.columns)
    for col_idx, header in enumerate(headers_sections, start=2):
        cell = ws.cell(row=current_row, column=col_idx, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    
    current_row += 1
    
    # Données des sections
    for row_data in df_sections.itertuples(index=False):
        for col_idx, value in enumerate(row_data, start=2):
            cell = ws.cell(row=current_row, column=col_idx, value=value)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
        current_row += 1
    
    # Ajuster la largeur des colonnes
    for col in range(2, 17):  # Colonnes B à P
        ws.column_dimensions[chr(64 + col)].width = 15
    
    # Sauvegarder le fichier
    wb.save(nom_fichier)
    wb.close()
    
    print(f" Fichier Excel '{nom_fichier}' créé avec succès !")


# ========== TEST ==========

if __name__ == "__main__":
    
    # Données des matériaux
    data_material = {
        1: Material(
            name='S355',
            temperature=50,
            E=np.float64(200000.0),
            Sy=np.float64(312.0),
            Su=np.float64(470.0),
            poisson=np.float64(0.3)
        ),
        1001: Material(
            name='S355',
            temperature=50,
            E=np.float64(185000.0),
            Sy=np.float64(206.0),
            Su=np.float64(470.0),
            poisson=np.float64(0.3)
        )
    }

    # Données des membres
    data_beam = {
        1: Element(
            id=1,
            nodes_id=[1, 2, 3, 4],
            section="UPA100",
            material="S235",
            lambda_rccm=100, 
            Lb=100
        ),
        2: Element(
            id=3,
            nodes_id=[31, 32, 33, 34],
            section="HEB120",
            material="S275",
            lambda_rccm=2000, 
            Lb=1000
        ),
    }
    
    # Données des sections - Test manuel comme Material et Element
    data_sections = {
        1: Section(
            name='IPE 100',
            is_closed=False,
            h=100.0,
            l=55.0,
            D=None,
            tw=4.1,
            tf=5.7,
            A=1030.0,
            Iy=1710000.0,
            Iz=159000.0,
            ry=40.7,
            rz=12.4,
            Am=515.0,
            b_t=4.8,
            d_t=21.5,
            Sp=[
                Section.StressPoint(
                    sec_name='IPE 100',
                    id=1,
                    coordinates=(50.0, 27.5),
                    Qy=19700.0,
                    Qz=2155.3,
                    e=5.7,
                    Wno=0.0,
                    Sw=0.0
                )
            ]
        ),
        2: Section(
            name='HEB 120',
            is_closed=False,
            h=120.0,
            l=120.0,
            D=None,
            tw=6.5,
            tf=11.0,
            A=3400.0,
            Iy=8640000.0,
            Iz=2890000.0,
            ry=50.4,
            rz=29.2,
            Am=1700.0,
            b_t=5.5,
            d_t=15.1,
            Sp=[
                Section.StressPoint(
                    sec_name='HEB 120',
                    id=1,
                    coordinates=(60.0, 60.0),
                    Qy=86400.0,
                    Qz=34200.0,
                    e=11.0,
                    Wno=0.0,
                    Sw=0.0
                )
            ]
        )
    }
    
    # Combinaisons (optionnel)
    combinations_data = {
        'Situation 1': ['NP', 'CO3', 'CO4', 'CO5'],
        'Situation 2': ['ACC', 'CO9', 'CO10', 'CO11']
    }
    
    try:
        create_input(
            materials=data_material, 
            member=data_beam,
            sections=data_sections,
            combinations=combinations_data
        )
    except Exception as e:
        print(f" Erreur lors de la création : {e}")
        import traceback
        traceback.print_exc()